import os
import threading
from flask import Flask, request, jsonify, render_template, redirect, url_for, session, send_file
from supabase import create_client, Client
from werkzeug.utils import secure_filename
from datetime import datetime
import pytz

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Deklarasi Top-Level Wajib untuk Vercel Serverless
app = Flask(__name__)
app.secret_key = "kunci_rahasia_absensi_khoerul"

# Menjadikan handler global siap dibaca WSGI Vercel
app = app

# --- KONFIGURASI UTAMA DATABASE SUPABASE ---
SUPABASE_URL = "https://filpxlbzeallnqawjqbe.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImZpbHB4bGJ6ZWFsbG5xYXdqcWJlIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzkyNzYzODAsImV4cCI6MjA5NDg1MjM4MH0.XoGPWrU4qg4DyI1g6A7pAKzkV0pFyfCosYP645JKY7A"
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

UPLOAD_FOLDER = 'static/uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# ========================================================
# 2. HELPER FUNCTIONS: FORMAT DATA, RELASI, & KONVERSI WIB
# ========================================================

def get_mapped_students():
    try:
        res = supabase.table("students").select("*").order("id", desc=False).execute()
        return {item['id']: item for item in res.data}
    except Exception as e:
        print(f"[ERROR] Gagal memetakan data siswa: {str(e)}")
        return {}

def format_to_wib(iso_timestamp_str):
    try:
        if not iso_timestamp_str:
            return "-"
        clean_timestamp = iso_timestamp_str.replace("Z", "+00:00")
        utc_dt = datetime.fromisoformat(clean_timestamp)
        wib_timezone = pytz.timezone("Asia/Jakarta")
        wib_dt = utc_dt.astimezone(wib_timezone)
        return wib_dt.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return iso_timestamp_str

def format_indonesian_month(date_str):
    try:
        month_num = date_str.split("/")[1]
        months = {
            "01": "Januari", "02": "Februari", "03": "Maret", "04": "April",
            "05": "Mei", "06": "Juni", "07": "Juli", "08": "Agustus",
            "09": "September", "10": "Oktober", "11": "November", "12": "Desember"
        }
        return months.get(month_num, "Lainnya")
    except:
        return "Lainnya"

def get_hari_indonesia(wib_dt):
    hari_map = {
        "Monday": "Senin", "Tuesday": "Selasa", "Wednesday": "Rabu",
        "Thursday": "Kamis", "Friday": "Jumat", "Saturday": "Sabtu", "Sunday": "Minggu"
    }
    return hari_map.get(wib_dt.strftime("%A"), "Senin")

def save_to_supabase_background(student_id, photo_url):
    try:
        supabase.table("attendance").insert({
            "student_id": int(student_id),
            "photo_url": photo_url,
            "status": "Hadir"
        }).execute()
    except Exception as e:
        print(f"[Supabase Error] Gagal sinkronisasi: {str(e)}")


# ========================================================
# 3. API GATEWAY: MENERIMA DATA DARI ESP32-CAM
# ========================================================

@app.route('/api/attendance', methods=['POST'])
def receive_attendance():
    try:
        student_id = request.form.get('student_id')
        file = request.files.get('photo')

        if not student_id or not file:
            return jsonify({"status": "failed", "message": "Missing data"}), 400

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = secure_filename(f"absen_{student_id}_{timestamp}.jpg")
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        file.save(file_path)
        photo_url = f"/static/uploads/{filename}"

        threading.Thread(target=save_to_supabase_background, args=(student_id, photo_url)).start()
        return jsonify({"status": "success", "message": "Buffered instantly"}), 201
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


# ========================================================
# 4. PORTAL MONITORING ORANG TUA
# ========================================================

@app.route('/orangtua/login', methods=['GET', 'POST'])
def orangtua_login():
    if request.method == 'POST':
        student_id = request.form.get('student_id')
        parent_phone = request.form.get('parent_phone')
        try:
            user_check = supabase.table("students").select("*").eq("id", int(student_id)).eq("parent_phone", parent_phone).execute()
            if len(user_check.data) > 0:
                session['logged_in_parent'] = True
                session['student_id'] = int(student_id)
                return redirect(url_for('orangtua_dashboard'))
            else:
                return render_template('login_orangtua.html', error="ID atau No. HP Salah!")
        except Exception:
            return render_template('login_orangtua.html', error="Koneksi database gagal!")
    return render_template('login_orangtua.html')

@app.route('/orangtua')
def orangtua_dashboard():
    if not session.get('logged_in_parent'): 
        return redirect(url_for('orangtua_login'))
    my_student_id = session.get('student_id')
    try:
        student_profile_res = supabase.table("students").select("*").eq("id", my_student_id).single().execute()
        student_profile = student_profile_res.data if student_profile_res.data else {}
        return render_template('orangtua.html', student_id=my_student_id, parent_name=student_profile.get('parent_name', '-'), student_name=student_profile.get('name', '-'), class_level=student_profile.get('class_level', '-'), major=student_profile.get('major', '-'))
    except Exception as e: 
        return f"Gagal Memuat Dashboard Orangtua: {str(e)}", 500

@app.route('/api/orangtua/data/<int:student_id>')
def api_orangtua_data(student_id):
    try:
        hours_res = supabase.table("school_hours").select("*").execute().data
        hours_map = {h['day_name']: h for h in hours_res}

        attendance_query = supabase.table("attendance").select("*").eq("student_id", student_id).order("id", desc=True).execute().data
        permissions_query = supabase.table("permissions").select("*").eq("student_id", student_id).order("permission_date", desc=False).execute().data
        
        tz_jkt = pytz.timezone("Asia/Jakarta")
        now_jkt = datetime.now(tz_jkt)
        today_str = now_jkt.strftime("%d/%m/%Y")
        hari_ini = get_hari_indonesia(now_jkt)

        config_hari_ini = hours_map.get(hari_ini, {"time_in": "07:00:00", "time_out": "14:00:00"})

        my_child_logs = []
        detailed_recap = {}
        today_logs = []

        for log in attendance_query:
            raw_time = log.get('created_at', '')
            wib_time = format_to_wib(raw_time)
            log['created_at'] = wib_time
            
            clean_timestamp = raw_time.replace("Z", "+00:00")
            dt_obj = datetime.fromisoformat(clean_timestamp).astimezone(tz_jkt)
            hari_log = get_hari_indonesia(dt_obj)
            config_log = hours_map.get(hari_log, {"time_in": "07:00:00", "time_out": "14:00:00"})
            
            if log.get('status') == 'Hadir':
                jam_menit_log = dt_obj.strftime("%H:%M:%S")
                if jam_menit_log > config_log['time_in']:
                    log['status'] = "Kesiangan"
            
            my_child_logs.append(log)
            if wib_time.startswith(today_str):
                today_logs.append(log)
            
            nama_bulan = format_indonesian_month(wib_time)
            tanggal_baca = wib_time.split(" ")[0]
            if nama_bulan not in detailed_recap:
                detailed_recap[nama_bulan] = []
            detailed_recap[nama_bulan].append({"tanggal": tanggal_baca, "status": log['status']})

        today_logs.reverse()
        datang_info = {"waktu": "--:--", "foto": "", "status": "Tidak Hadir"}
        pulang_info = {"waktu": "--:--", "foto": "", "status": "Bolos"}
        
        if config_hari_ini['time_in'] == "00:00:00":
            datang_info["status"] = "Libur Sekolah"
            pulang_info["status"] = "Libur Sekolah"
        else:
            if len(today_logs) > 0:
                jam_datang = today_logs[0]['created_at'].split(" ")[1]
                datang_info["waktu"] = jam_datang
                datang_info["foto"] = today_logs[0]['photo_url']
                datang_info["status"] = "Kesiangan" if jam_datang > config_hari_ini['time_in'] else "Tepat Waktu"
                
            if len(today_logs) > 1:
                pulang_info["waktu"] = today_logs[-1]['created_at'].split(" ")[1]
                pulang_info["foto"] = today_logs[-1]['photo_url']
                pulang_info["status"] = "Sudah Pulang"
            elif len(today_logs) == 1:
                jam_sekarang = now_jkt.strftime("%H:%M:%S")
                if jam_sekarang > config_hari_ini['time_out']:
                    pulang_info["status"] = "Bolos"
                else:
                    pulang_info["status"] = "Berada di Sekolah"

        return jsonify({
            "my_child_logs": my_child_logs,
            "my_permission_logs": permissions_query,
            "detailed_recap": detailed_recap,
            "quick_access": {
                "tanggal": today_str,
                "hari": hari_ini,
                "jam_masuk_target": config_hari_ini['time_in'][:5],
                "jam_pulang_target": config_hari_ini['time_out'][:5],
                "datang": datang_info,
                "pulang": pulang_info
            }
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/orangtua/izin', methods=['POST'])
def orangtua_submit_izin():
    if not session.get('logged_in_parent'): 
        return redirect(url_for('orangtua_login'))
    try:
        student_id = request.form.get('student_id')
        reason = request.form.get('reason')
        permission_date = request.form.get('date')
        file = request.files.get('voice_note')

        if file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = secure_filename(f"vn_{student_id}_{timestamp}.mp3")
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            proof_url = f"/static/uploads/{filename}"
        else:
            proof_url = ""

        supabase.table("permissions").insert({
            "student_id": int(student_id), "reason": reason, "permission_date": permission_date, "proof_url": proof_url, "status": "Pending"
        }).execute()
        return redirect(url_for('orangtua_dashboard'))
    except Exception as e:
        return f"Gagal mengirim izin: {str(e)}", 500

@app.route('/orangtua/logout')
def orangtua_logout():
    session.clear()
    return redirect(url_for('orangtua_login'))


# ========================================================
# 5. PORTAL UTAMA ADMIN SEKOLAH & DATA ENGINE (CRUD)
# ========================================================

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        if request.form.get('username') == "admin" and request.form.get('password') == "admin123":
            session['logged_in_admin'] = True
            return redirect(url_for('admin_dashboard'))
        else:
            return render_template('login_admin.html', error="Username/Password Salah!")
    return render_template('login_admin.html')

@app.route('/admin')
def admin_dashboard():
    if not session.get('logged_in_admin'): 
        return redirect(url_for('admin_login'))
    return render_template('admin.html')

@app.route('/api/admin/data')
def api_admin_data():
    try:
        students_map = get_mapped_students()
        attendance_query = supabase.table("attendance").select("*").order("id", desc=True).execute().data
        permissions_query = supabase.table("permissions").select("*").order("id", desc=True).execute().data
        hours_query = supabase.table("school_hours").select("*").execute().data
        
        attendance_logs = []
        for log in attendance_query:
            st_id = log.get('student_id')
            log['students'] = students_map.get(st_id, {"name": "Siswa", "class_level": "X", "major": "TKJ", "parent_name": "-"})
            log['created_at'] = format_to_wib(log.get('created_at', ''))
            attendance_logs.append(log)

        permissions_logs = []
        for perm in permissions_query:
            st_id = perm.get('student_id')
            perm['students'] = students_map.get(st_id, {"name": "Siswa", "class_level": "X", "major": "TKJ", "parent_name": "-"})
            permissions_logs.append(perm)

        return jsonify({
            "attendance_logs": attendance_logs,
            "permissions": permissions_logs,
            "students": list(students_map.values()),
            "school_hours": hours_query
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/admin/hours/update', methods=['POST'])
def api_update_school_hours():
    if not session.get('logged_in_admin'): 
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
    try:
        data = request.json
        day = data.get('day_name')
        time_in = data.get('time_in')
        time_out = data.get('time_out')
        
        formatted_in = time_in if len(time_in) == 8 else f"{time_in}:00"
        formatted_out = time_out if len(time_out) == 8 else f"{time_out}:00"
        
        supabase.table("school_hours").update({
            "time_in": formatted_in,
            "time_out": formatted_out
        }).eq("day_name", day).execute()
        
        return jsonify({"status": "success", "message": f"Waktu hari {day} sukses diperbarui!"}), 200
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/admin/laporan/download-excel', methods=['GET'])
def download_laporan_excel_formal():
    if not session.get('logged_in_admin'): 
        return "Unauthorized", 401
    try:
        f_major = request.args.get('major', 'ALL')
        f_name = request.args.get('name', '').lower()
        f_date = request.args.get('date', '') 
        f_month = request.args.get('month', 'ALL') 

        fmt_date = ""
        if f_date:
            p = f_date.split("-")
            fmt_date = f"{p[2]}/{p[1]}/{p[0]}"

        students_map = get_mapped_students()
        attendance_logs = supabase.table("attendance").select("*").order("id", desc=True).execute().data

        wb = openpyxl.Workbook()
        ws_sum = wb.active
        ws_sum.title = "Ringkasan Absensi"
        ws_det = wb.create_sheet(title="Log Detail Real-time")

        ws_sum.views.sheetView[0].showGridLines = True
        ws_det.views.sheetView[0].showGridLines = True

        font_title = Font(name="Arial", size=13, bold=True, color="1E293B")
        font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        font_body = Font(name="Arial", size=10)
        
        fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        fill_yellow = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
        fill_red = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        thin_border = Border(left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'), top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1'))

        ws_sum.append(["PEMERINTAH PROVINSI JAWA TENGAH"])
        ws_sum.append(["DINAS PENDIDIKAN DAN KEBUDAYAAN"])
        ws_sum.append(["SMK NEGERI MONITORING ELEKTRONIK BIOMETRIK"])
        ws_sum.append(["LAPORAN REKAPITULASI CAPAIAN KEHADIRAN SISWA"])
        ws_sum.append([])
        for r in range(1, 5): ws_sum.cell(row=r, column=1).font = font_title

        headers_sum = ["ID SISWA", "NAMA LENGKAP SISWA", "KELAS", "JURUSAN", "HADIR (HARI)", "TELAT (HARI)", "BOLOS (HARI)", "RATA-RATA %"]
        ws_sum.append(headers_sum)
        for c, text in enumerate(headers_sum, 1):
            cell = ws_sum.cell(row=6, column=c)
            cell.font = font_header; cell.fill = fill_header; cell.alignment = align_center; cell.border = thin_border

        ws_det.append(["LOG REAL-TIME PRESENSI MASUK & PULANG HARDWARE IOT"])
        ws_det.append([f"Ekstraksi filter harian otomatis tanggal cetak: {datetime.now().strftime('%d/%m/%Y')}"])
        ws_det.append([])
        ws_det.cell(row=1, column=1).font = font_title; ws_det.cell(row=2, column=1).font = Font(name="Arial", size=10, italic=True, color="64748B")

        headers_det = ["ID SISWA", "NAMA LENGKAP SISWA", "TINGKAT & JURUSAN", "TANGGAL ACESS", "JAM ABSEN", "STATUS"]
        ws_det.append(headers_det)
        for c, text in enumerate(headers_det, 1):
            cell = ws_det.cell(row=4, column=c)
            cell.font = font_header; cell.fill = fill_header; cell.alignment = align_center; cell.border = thin_border

        row_sum_idx = 7; row_det_idx = 5

        for st_id, s in students_map.items():
            if f_major != "ALL" and s['major'] != f_major: continue
            if f_name and f_name != "" and f_name not in s['name'].lower(): continue

            cHadir = 0; cTelat = 0; cBolos = 0
            for log in attendance_logs:
                if log['student_id'] == s['id']:
                    wib_t = format_to_wib(log['created_at'])
                    l_date = wib_t.split(" ")[0]
                    l_month = l_date.split("/")[1]

                    if f_date and l_date != fmt_date: continue
                    if f_month != "ALL" and l_month != f_month: continue

                    if log['status'] == "Hadir": cHadir += 1
                    elif log['status'] == "Kesiangan": cTelat += 1
                    elif log['status'] == "Bolos": cBolos += 1

                    ws_det.append([s['id'], s['name'], f"{s['class_level']} {s['major']}", l_date, wib_t.split(" ")[1], log['status']])
                    for col_d in range(1, 7):
                        cell_d = ws_det.cell(row=row_det_idx, column=col_d)
                        cell_d.font = font_body; cell_d.border = thin_border; cell_d.alignment = align_center if col_d != 2 else align_left
                        if col_d == 6:
                            if log['status'] == "Kesiangan": cell_d.fill = fill_yellow
                            elif log['status'] == "Bolos": cell_d.fill = fill_red
                    row_det_idx += 1

            formula_pct = f"=ROUND((E{row_sum_idx}+F{row_sum_idx})/(E{row_sum_idx}+F{row_sum_idx}+G{row_sum_idx})*100,0)&\"%\""
            ws_sum.append([s['id'], s['name'], s['class_level'], s['major'], cHadir, cTelat, cBolos, formula_pct])
            for col_s in range(1, 9):
                cell_s = ws_sum.cell(row=row_sum_idx, column=col_s)
                cell_s.font = font_body; cell_s.border = thin_border; cell_s.alignment = align_center if col_s != 2 else align_left
                if row_sum_idx % 2 == 0: cell_s.fill = fill_zebra
            row_sum_idx += 1

        for ws in [ws_sum, ws_det]:
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

        output_path = "static/uploads/Laporan_Presensi_SMK_Formal.xlsx"
        wb.save(output_path)
        return send_file(output_path, as_attachment=True, download_name=f"Laporan_Presensi_SMK.xlsx")
    except Exception as e:
        return f"Gagal: {str(e)}", 500

@app.route('/admin/permissions/quick-verify/<int:perm_id>/<string:new_status>', methods=['GET'])
def quick_verify_permission(perm_id, new_status):
    if not session.get('logged_in_admin'): return redirect(url_for('admin_login'))
    try:
        perm_res = supabase.table("permissions").select("*").eq("id", perm_id).single().execute().data
        if not perm_res: return "Gagal", 404
        student_id = perm_res.get('student_id'); p_date = perm_res.get('permission_date')
        reason = perm_res.get('reason', 'Izin'); proof_url = perm_res.get('proof_url', '')

        supabase.table("permissions").update({"status": new_status}).eq("id", perm_id).execute()
        if new_status == "Approve" and student_id and p_date:
            start_day = f"{p_date}T00:00:00+00:00"; end_day = f"{p_date}T23:59:59+00:00"
            check_exist = supabase.table("attendance").select("id").eq("student_id", int(student_id)).gte("created_at", start_day).lte("created_at", end_day).execute().data
            if len(check_exist) == 0:
                supabase.table("attendance").insert({"student_id": int(student_id), "status": reason, "photo_url": proof_url}).execute()
        return redirect(url_for('admin_dashboard'))
    except Exception as e: return str(e), 500

# Endpoint API Penengah untuk Fitur Perintah Jari IoT (Enroll/Delete)
@app.route('/api/admin/fingerprint/command', methods=['POST'])
def send_fingerprint_command():
    if not session.get('logged_in_admin'):
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
    try:
        data = request.json
        student_id = int(data.get('student_id'))
        action = data.get('action')
        
        supabase.table("system_commands").insert({
            "student_id": student_id,
            "command": action,
            "status": "PENDING"
        }).execute()
        return jsonify({"status": "success", "message": f"Sinyal {action} berhasil dipancarkan ke hardware IoT!"}), 200
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/admin/update-student/<int:id>', methods=['POST'])
def update_student(id):
    if not session.get('logged_in_admin'): return redirect(url_for('admin_login'))
    try:
        supabase.table("students").update({"name": request.form.get('name'), "parent_name": request.form.get('parent_name'), "class_level": request.form.get('class_level'), "major": request.form.get('major'), "parent_phone": request.form.get('parent_phone')}).eq("id", id).execute()
        return redirect(url_for('admin_dashboard'))
    except Exception as e: return str(e), 500

@app.route('/admin/add-student', methods=['POST'])
def add_student():
    if not session.get('logged_in_admin'): return redirect(url_for('admin_login'))
    try:
        supabase.table("students").insert({"id": int(request.form.get('student_id')), "name": request.form.get('name'), "parent_name": request.form.get('parent_name'), "parent_phone": request.form.get('parent_phone'), "class_level": request.form.get('class_level'), "major": request.form.get('major')}).execute()
        return redirect(url_for('admin_dashboard'))
    except Exception as e: return str(e), 500

@app.route('/admin/logout')
def admin_logout():
    session.clear()
    return redirect(url_for('admin_login'))

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
